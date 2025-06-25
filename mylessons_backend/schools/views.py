import json
import datetime as dt
import math
from decimal import Decimal, InvalidOperation

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from django.shortcuts import render, get_object_or_404
from locations.models import Location
from equipment.serializers import EquipmentSerializer
from schools.serializers import CSVUploadSerializer, UpdateContactsSerializer, ContactsSerializer
from sports.models import Sport
from payments.models import Payment
from users.models import Instructor, Monitor, Student, UserAccount
from .models import Review, School
from datetime import datetime, timedelta
from django.contrib.auth import authenticate, get_user_model
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status, views, serializers, permissions
from django.views.decorators.csrf import csrf_exempt
from rest_framework.permissions import AllowAny, IsAuthenticated
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests
import logging
from django.contrib.auth.hashers import make_password
from notifications.models import Notification
from lessons.models import Lesson, Pack
from events.models import Activity
from schools.models import School
from django.db.models import Q, Sum
from django.utils.timezone import now
from django.http import JsonResponse
from django.db import transaction
import pandas as pd
from django.http import HttpResponse
import io
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule

logger = logging.getLogger(__name__)


class UpdateContactsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, school_id):
        school = get_object_or_404(School, id=school_id)

        serializer = UpdateContactsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # serializer.validated_data['contacts'] is a dict {'teams': [...]}
        school.contacts = serializer.validated_data['contacts']
        school.save(update_fields=['contacts'])

        # Optionally return the updated contacts
        return Response(
            ContactsSerializer(school.contacts).data,
            status=status.HTTP_200_OK
        )

@permission_classes([IsAuthenticated])
@api_view(['POST'])
def get_school_time_limit(request):
    """
    Expects a POST with JSON data containing "school_name".
    Returns the schedule_time_limit for the given school.
    """
    school_name = request.data.get('school_name')
    if not school_name:
        return Response(
            {'error': "Missing 'school_name' parameter."},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        school = School.objects.get(name=school_name)
    except School.DoesNotExist:
        logger.debug("School not found for school name %s", school_name)
        return Response(
            {'error': "School not found."},
            status=status.HTTP_404_NOT_FOUND
        )

    return Response(
        {'time_limit': school.schedule_time_limit},
        status=status.HTTP_200_OK
    )

@permission_classes([IsAuthenticated])
@api_view(['GET'])
def get_services(request, school_id):
    """
    GET /api/schools/<school_id>/services/
    Returns the list of services for the specified school.
    """
    school = get_object_or_404(School, pk=school_id)
    return Response(school.services, status=200)

@permission_classes([IsAuthenticated])
@api_view(['GET'])
def get_equipments(request, school_id, subject_id):
    """
    GET /api/schools/<school_id>/equipments/?sport=<sport_id>
    Returns the list of equipments for the specified school,
    optionally filtered to only those used in a given sport.
    """
    school = get_object_or_404(School, pk=school_id)
    qs = school.equipments.all()

    if subject_id:
        qs = qs.filter(sports__id=subject_id)

    serializer = EquipmentSerializer(qs, many=True)
    return Response(serializer.data, status=200)

@permission_classes([IsAuthenticated])
@api_view(['POST'])
def add_edit_service(request, school_id):
    """
    POST /api/schools/<school_id>/services/add_edit/
    
    Request Body (JSON):
    [
    {
        "id": "cc6ab39c-b01a-46f8-938b-41753ec62036",
        "name": "Private Lessons",
        "type": {
        "pack": "private"
        },
        "photos": [],
        "sports": [],
        "details": {
        "pricing_options": [
            {
            "price": 200.0,
            "people": 1,
            "classes": 8,
            "duration": 60,
            "time_limit": 90
            }
        ]
        },
        "benefits": [],
        "currency": "EUR",
        "locations": [],
        "description": ""
    }
    ]


    If a service with the given 'id' exists, it will be updated.
    Otherwise, a new service is appended.

    Returns the updated list of services.
    """

    # TODO checks for conflicts with missing payment types, for example a service for private lessons with pricing option for 60min 4 people and not payment types for all the roles for those details 

    school = get_object_or_404(School, pk=school_id)
    service_data = request.data  # JSON payload

    try:
        updated_services = school.add_or_edit_service(service_data)
    except ValueError as e:
        return Response({"detail": str(e)}, status=400)

    return Response(updated_services, status=200)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def all_schools(request):
    schools = School.objects.all()
    data = []
    for school in schools:
        if school.logo and hasattr(school.logo, 'url'):
            logo_url = request.build_absolute_uri(school.logo.url)
        else:
            logo_url = ""
            
        data.append({
            'school_id': school.id,
            'school_name': school.name,
            'list_of_locations': [
                {'id': location.id, 'name': location.name}
                for location in school.locations.all()
            ],
            'isFavorite': True if request.user.schools.exists() and school in request.user.schools.all() else False,
            'services': school.services,
            'currency': school.currency if school.currency else "EUR",
            'contacts': school.contacts,
            'image' : logo_url,
            'subjects': [{
                'id': subject.id,
                'name': subject.name,
                'locations': [{
                    'id': location.id,
                    'name': location.name,
                    'address': location.address if location.address else "",
                    'instructors': [{
                        'id': instructor.id,
                        'name': str(instructor),
                    } for instructor in subject.instructors.filter(id__in=location.instructors.all())],
                } for location in school.locations.all()]
            } for subject in school.sports.all()]
        })
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def school_details_view(request):
    user = request.user
    try:
        school = School.objects.get(pk=user.current_school_id)
    except School.DoesNotExist:
        return Response({
            'success': True,
            'school_id': None,
            'school_name': "",
            'services': {},
            'payment_types': {},
            'staff': [],
            'currency': "",
            'locations': [],
            'subjects': [],
            'equipments': [],
            'critical_message': ""  # no conflicts
        })

    # Use a dict to merge staff by user id
    staff_dict = {}

    # Process Admins
    for admin in school.admins.all():
        # assuming admin is a UserAccount instance
        u = admin
        if u.id not in staff_dict:
            staff_dict[u.id] = {
                'user_id': u.id,
                'user_name': f"{u.first_name} {u.last_name}",
                'roles': [],
                'payment_types': u.payment_types[school.name] if school.name in u.payment_types else {},  # full payment_types dict
            }
        staff_dict[u.id]['roles'].append("Admin")

    # Process Instructors (access user via instructor.user)
    for instructor in school.instructors.all():
        u = instructor.user
        if u.id not in staff_dict:
            staff_dict[u.id] = {
                'user_id': u.id,
                'user_name': f"{u.first_name} {u.last_name}",
                'roles': [],
                'payment_types': u.payment_types[school.name],
            }
        staff_dict[u.id]['roles'].append("Instructor")

    # Process Monitors (access user via monitor.user)
    for monitor in school.monitors.all():
        u = monitor.user
        if u.id not in staff_dict:
            staff_dict[u.id] = {
                'user_id': u.id,
                'user_name': f"{u.first_name} {u.last_name}",
                'roles': [],
                'payment_types': u.payment_types[school.name],
            }
        staff_dict[u.id]['roles'].append("Monitor")

    staff_list = list(staff_dict.values())

    # Call the conflict verification method.
    conflict_message = school.check_payment_types_conflicts()
    
    equipment_list = []
    for eq in school.equipments.all():
        equipment_list.append({
            'equipment_id':      eq.id,
            'equipment_name':    eq.name,
            'location_id':       eq.location.id if eq.location else None,
            'location_name':     eq.location.name if eq.location else "",
            'state':             eq.state or "",
            'photo_url':         request.build_absolute_uri(eq.photo.url)
                                   if eq.photo else "",
            'size':              eq.size or "",
            'is_for_kids':       eq.is_for_kids,
            'description':       eq.description or "",
            'brand':             eq.brand or "",
            'subjects': [
                {'id': s.id, 'name': s.name}
                for s in eq.sports.all()
            ],
        })


    return Response({
        'success': True,
        'school_id': school.id,
        'school_name': school.name,
        'services': school.services,
        'payment_types': school.payment_types,
        'staff': staff_list,
        'currency': school.currency,
        'locations': [
            {
                'location_id': loc.id,
                'location_name': loc.name,
                'address': loc.address or "",
            }
            for loc in school.locations.all()
        ],
        'subjects': [
            {
                'subject_id': subj.id,
                'subject_name': subj.name,
            }
            for subj in school.sports.all()
        ],
        'equipment': equipment_list,
        'critical_message': conflict_message,
    })



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_payment_type_view(request):
    """
    Update the default payment types for the school, not a specific user
    
    """

    # TODO check for conflicts within the payment types, example an instructor payment type for private lessons for 60 min 1 to 4 people and an other for private lessons 60 min 1 to 2 people, if so, asks to override or edit
    # TODO it has to go through the school services and checks if there are any conflicts example having a service and not having a payment type to cover for that service details
    
    try:
        data = json.loads(request.body)
        logger.debug("Received payload: %s", data)
    except json.JSONDecodeError as e:
        logger.error("JSON decode error: %s", str(e))
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    
    school_id = data.get('school_id')
    school_name = data.get('school_name')
    user_id = data.get('user_id')
    user = request.user
    logger.debug("User: %s", user)
    
    if school_id:
        try:
            school = School.objects.get(id=school_id)
            logger.debug("Found school by id: %s", school)
        except School.DoesNotExist:
            logger.error("School not found for id: %s", school_id)
            return JsonResponse({'success': False, 'error': 'School not found'}, status=404)
    elif school_name:
        school = School.objects.create(name=school_name)
        school.admins.add(user)
        user.current_school_id = school.id
        user.save()
        logger.debug("Created school with name: %s", school_name)
    else:
        logger.error("Neither school_id nor school_name provided.")
        return JsonResponse({'success': False, 'error': 'Either school_id or school_name must be provided'}, status=400)
    
    key_path = data.get('key_path')
    new_value = data.get('new_value')
    logger.debug("key_path: %s, new_value: %s", key_path, new_value)
    
    if not key_path or new_value is None:
        logger.error("Missing key_path or new_value. key_path=%s, new_value=%s", key_path, new_value)
        return JsonResponse({'success': False, 'error': 'key_path and new_value must be provided'}, status=400)
    
    user_obj = None
    if user_id:
        try:
            user_obj = UserAccount.objects.get(id=user_id)  # or your user model
        except UserAccount.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
        
    try:
        result = school.update_payment_type_value(key_path, new_value, user_obj=user_obj)
        logger.debug("update_payment_type_value result: %s", result)
    except Exception as e:
        logger.exception("Error during update_payment_type_value:")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    logger.debug("Updated payment_types:\nschool: %s", school.payment_types)
    
    # Return updated data
    if user_obj:
        logger.debug("Updated payment_types:\nschool: %s", user_obj.payment_types)
        # If we updated a user’s payment_types
        return JsonResponse({
            'success': True,
            'school_id': school.id,
            'school_name': school.name,
            'user_id': user_obj.id,
            # user’s payment_types for this school
            'payment_types': user_obj.payment_types.get(school.name, {})
        })
    else:
        # If we updated the school’s default payment_types
        return JsonResponse({
            'success': True,
            'school_id': school.id,
            'school_name': school.name,
            'payment_types': school.payment_types
        })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_pack_price_view(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    
    # Retrieve the school by ID or create one if a school_name is provided.
    school_id = data.get('school_id')
    school_name = data.get('school_name')
    user = request.user
    
    if school_id:
        try:
            school = School.objects.get(id=school_id)
        except School.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'School not found'}, status=404)
    elif school_name:
        # Create a new school with the given name and add request.user to admins.
        school = School.objects.create(name=school_name)
        school.admins.add(user)
        user.current_school_id = school.id
        user.save()
        
    else:
        return JsonResponse({'success': False, 'error': 'Either school_id or school_name must be provided'}, status=400)
    
    # Extract pack price details from the request data.
    pack_type = data.get('pack_type')
    duration = data.get('duration')
    number_of_people = data.get('number_of_people')
    number_of_classes = data.get('number_of_classes')
    price = data.get('price')
    expiration_date = data.get('expiration_date')  # computed on the client based on time limit
    currency = data.get('currency')
    
    # Set the school's currency if provided.
    if currency:
        school.currency = currency
        school.save()

    try:
        school.update_pack_price(
            pack_type=pack_type,
            duration=duration,
            number_of_people=number_of_people,
            number_of_classes=number_of_classes,
            price=price,
            expiration_date=expiration_date
        )
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({
        'success': True,
        'school_id': school.id,
        'school_name': school.name,
        'pack_prices': school.pack_prices
    })
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def delete_payment_type_entry_view(request):
    try:
        data = json.loads(request.body)
        logger.debug("Received delete payload: %s", data)
    except json.JSONDecodeError as e:
        logger.error("JSON decode error: %s", str(e))
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    
    school_id = data.get('school_id')
    school_name = data.get('school_name')
    key_path = data.get('key_path')
    entry_to_delete = data.get('entry')
    user_id = data.get('user_id')
    
    if not school_id or not key_path or not entry_to_delete:
        return JsonResponse({'success': False, 'error': 'Missing required parameters'}, status=400)
    
    # Fetch school by id.
    try:
        school = School.objects.get(id=school_id)
    except School.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'School not found'}, status=404)
    
    # If user_id is provided, fetch the user.
    user_obj = None
    if user_id:
        try:
            user_obj = UserAccount.objects.get(id=user_id)
        except UserAccount.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
    
    try:
        result = school.delete_payment_type_value(key_path, entry_to_delete, user_obj=user_obj)
        logger.debug("delete_payment_type_value result: %s", result)
    except Exception as e:
        logger.exception("Error during delete_payment_type_value:")
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    # Return updated payment types.
    if user_obj:
        updated = user_obj.payment_types.get(school.name, {})
    else:
        updated = school.payment_types
    return JsonResponse({
        'success': True,
        'payment_types': updated
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def number_of_students_in_timeframe(request, school_id, start_date, end_date):

    # TODO 
    # 
    # account for students with lessons that dont have lessons scheduled
    #
    # maybe output the students ids
    #
    # TEST

    try:
        # Convert date strings to date objects
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

    user = request.user
    current_role = getattr(user, 'current_role', None)

    # Check if user has permission to view this school’s student count
    if current_role != "Admin" or school_id not in user.school_admins.values_list('id', flat=True):
        return Response({"error": "Not allowed to view this school's students data"}, status=403)

    lesson_students = Lesson.objects.filter(
        school_id=school_id
    ).filter(
        Q(date__gte=start_date, date__lte=end_date) |
        Q(date__isnull=True, pack__date__lte=end_date, pack__expiration_date__gte=start_date) |
        Q(date__isnull=True, pack__expiration_date__isnull=True, pack__date__lte=end_date)
    ).values_list('students', flat=True).distinct().count()

    activity_students = Activity.objects.filter(
        school_id=school_id,
        date__gte=start_date,
        date__lte=end_date
    ).values_list('students', flat=True).distinct().count()

    total_students = lesson_students + activity_students

    data = {"total_students": str(total_students)}
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def number_of_instructors_in_timeframe(request, school_id, start_date, end_date):

    # TODO TEST

    try:
        # Convert date strings to date objects
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

    user = request.user
    current_role = getattr(user, 'current_role', None)

    # Check if user has permission to view this school’s student count
    if current_role != "Admin" or school_id not in user.school_admins.values_list('id', flat=True):
        return Response({"error": "Not allowed to view this school's instructors data"}, status=403)

    # Count distinct students from lessons and activities in the given timeframe
    lesson_instructors = Lesson.objects.filter(
        school_id=school_id,
        date__gte=start_date,
        date__lte=end_date
    ).values_list('instructors', flat=True).distinct().count()

    activity_instructors = Activity.objects.filter(
        school_id=school_id,
        date__gte=start_date,
        date__lte=end_date
    ).values_list('instructors', flat=True).distinct().count()

    total_instructors = lesson_instructors + activity_instructors

    data = {"total_instructors": str(total_instructors)}
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def school_revenue_in_timeframe(request, school_id, start_date, end_date):
    
    # TODO TEST

    try:
        # Convert date strings to date objects
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

    user = request.user
    current_role = getattr(user, 'current_role', None)

    # Check if user has permission to view this school’s revenue
    if current_role != "Admin" or school_id not in user.school_admins.values_list('id', flat=True):
        return Response({"error": "Not allowed to view this school's revenue"}, status=403)

    # Calculate total revenue (sum of all payments linked to lessons in this school within the timeframe)
    total_revenue = Payment.objects.filter(
        school_id=school_id,
        date__gte=start_date,
        date__lte=end_date,
        instructor__isnull=True,  # Only payments where instructors are null
        monitor__isnull=True  # Only payments where monitors are null
    ).aggregate(total_income=Sum('value'))['total_income'] or 0  # Default to 0 if None

    data = {"total_revenue": str(total_revenue)}
    return Response(data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def number_of_bookings_in_timeframe(request, school_id, start_date, end_date):
    try:
        # Convert date strings to datetime objects
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

    user = request.user
    current_role = getattr(user, 'current_role', None)

    # Check if user has permission to view this school’s bookings count
    if current_role != "Admin" or school_id not in user.school_admins.values_list('id', flat=True):
        return Response({"error": "Not allowed to view this school's bookings"}, status=403)

    
    number_of_lessons = Lesson.objects.filter(
        school__id=school_id,
        date__gte=start_date,
        date__lte=end_date,
        start_time__isnull=False
    ).count()


    data = {"number_of_lessons_booked": str(number_of_lessons)}
    return Response(data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_instructor(request):
    """
    Permite aos administradores da escola adicionarem um instrutor
    """
    user = request.user
    current_role = user.current_role
    instructor_id = request.data.get("instructor_id")

    if not instructor_id :
        return Response({"error": "É necessário fornecer instructor_id"},
                        status=status.HTTP_400_BAD_REQUEST)
    
    if current_role == "Admin":
        instructor = Instructor.objects.get(id=instructor_id)
        school = School.objects.get(admins__in=[user])
        if school and instructor:
            success = school.add_instructor(instructor=instructor)
    else:
        return Response({"error": "Não tem permissão para adicionar o instrutor."},
                        status=status.HTTP_403_FORBIDDEN)
    if success:
        return Response({'message': 'Instrutor adicionado com sucesso!'}, status=status.HTTP_200_OK)
    else:
        return Response({"error": "Impossivel adicionar o instrutor."}, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def remove_instructor(request):
    """
    Permite aos administradores da escola adicionarem um instrutor
    """
    user = request.user
    current_role = user.current_role
    instructor_id = request.data.get("lesson_id")

    if not instructor_id :
        return Response({"error": "É necessário fornecer instructor_id"},
                        status=status.HTTP_400_BAD_REQUEST)
    
    if current_role == "Admin":
        instructor = Instructor.objects.filter(id=instructor_id)
        school = School.objects.get(admins__in=[user])
        if school and instructor:
            success = school.remove_instructor(instructor=instructor)
    else:
        return Response({"error": "Não tem permissão para remover o instrutor."},
                        status=status.HTTP_403_FORBIDDEN)
    if success:
        return Response({'message': 'Instrutor removido com sucesso!'}, status=status.HTTP_200_OK)
    else:
        return Response({"error": "Impossivel remover o instrutor."}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def create_school(request):
    # Read name from either JSON body or form-data
    school_name = request.data.get('school_name', '').strip()
    if not school_name:
        return Response(
            {'success': False, 'error': 'School name is required.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    user = request.user
    if user.current_role != "Admin":
        return Response(
            {'success': False, 'error': 'Only admins can create a school.'},
            status=status.HTTP_403_FORBIDDEN
        )

    if School.objects.filter(name__iexact=school_name).exists():
        return Response(
            {'success': False, 'error': 'School name already chosen.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Create the new school (without logo for now)
    school = School.objects.create(name=school_name)
    school.admins.add(user)
    user.current_school_id = school.id
    user.save()

    # Handle optional image upload
    image_file = request.FILES.get('image')
    if image_file:
        school.logo.save(image_file.name, image_file, save=True)

    # Build response payload
    school_data = {
        'id': school.id,
        'name': school.name,
    }
    if getattr(school, 'logo', None):
        # full URL to the logo, assuming MEDIA_URL is configured
        logo_url = request.build_absolute_uri(school.logo.url)
        school_data['logo'] = logo_url

    return Response(
        {'success': True, 'school': school_data},
        status=status.HTTP_201_CREATED
    )
    

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def check_user_view(request):
    """
    POST data: { "email": "somebody@example.com" }

    Returns the user's email, first_name, last_name if found.
    """
    email = request.data.get("email", "").strip()
    if not email:
        return Response({
            "success": False,
            "detail": "Email is required."
        }, status=400)

    try:
        user = UserAccount.objects.get(email=email)
        return Response({
            "success": True,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name
        }, status=200)
    except UserAccount.DoesNotExist:
        return Response({
            "success": False,
            "detail": "User not found."
        }, status=404)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_staff_view(request):
    """
    POST data:
    {
      "email": "somebody@example.com",
      "roles": ["Admin", "Instructor", "Monitor"]
    }

    Adds the user to the given roles in the current school.
    """
    email = request.data.get("email", "").strip()
    roles = request.data.get("roles", [])

    if not email:
        return Response({
            "success": False,
            "detail": "Email is required."
        }, status=400)

    # Attempt to fetch the user by email
    try:
        user = UserAccount.objects.get(email=email)
    except UserAccount.DoesNotExist:
        return Response({
            "success": False,
            "detail": "User not found."
        }, status=404)

    # Get the school associated with the current user
    try:
        school = School.objects.get(pk=request.user.current_school_id)
    except School.DoesNotExist:
        return Response({
            "success": False,
            "detail": "Current school not found for this user."
        }, status=404)

    # For each role, add the user to the appropriate relationship
    for role in roles:
        lower_role = role.lower()
        if lower_role == "admin":
            # School has a ManyToManyField or similar for admins
            school.admins.add(user)

        elif lower_role == "instructor":
            # Usually we have an Instructor model that references UserAccount
            instructor, _ = Instructor.objects.get_or_create(user=user)
            school.instructors.add(instructor)

        elif lower_role == "monitor":
            monitor, _ = Monitor.objects.get_or_create(user=user)
            school.monitors.add(monitor)
        # else ignore any unknown roles, or handle them as you see fit
        
        school.add_payment_types_to_user(user)

    return Response({
        "success": True,
        "detail": f"User {email} added to roles: {roles} in school {school.name}"
    }, status=200)
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_subjects(request):
    """
    Returns a list of all subjects (Sports) as dictionaries.
    If a school_id is provided, returns subjects linked to that school (school.sports.all()).
    If a lesson_id or pack_id is provided, also returns the selected subject (lesson.sport or pack.sport).
    """
    lesson_id = request.GET.get('lesson_id')
    pack_id = request.GET.get('pack_id')
    school_id = request.GET.get('school_id')

    selected_subject = None
    selected_subjects = None
    subjects_qs = Sport.objects.all().values()

    # If school_id provided, get subjects from that school; otherwise, get all subjects.
        

    if school_id:
        try:
            school = School.objects.get(pk=school_id)
            selected_subjects = [
                {
                    "id": subject.id,
                    "name": subject.name,
                }
                for subject in school.sports.all()
            ]
            
        except School.DoesNotExist:
            print("Error on school_id")
    # Check for lesson_id or pack_id to set the selected subject.
    elif lesson_id:
        try:
            lesson = Lesson.objects.get(pk=lesson_id)
            selected_subject_obj = lesson.sport  # assuming lesson.sport is a Sport instance
            
            selected_subject = {"id": selected_subject_obj.id, "name": selected_subject_obj.name} if selected_subject_obj else None
        except Lesson.DoesNotExist:
            selected_subject = None
    elif pack_id:
        try:
            pack = Pack.objects.get(pk=pack_id)
            selected_subject_obj = pack.sport  # assuming pack.sport is a Sport instance
            selected_subject = {"id": selected_subject_obj.id, "name": selected_subject_obj.name} if selected_subject_obj else None
        except Pack.DoesNotExist:
            selected_subject = None

    response_data = {
        "subjects": list(subjects_qs),
    }
    if selected_subject:
        response_data["selected_subject"] = selected_subject
    if selected_subjects:
        response_data["selected_subjects"] = selected_subjects

    return Response(response_data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_locations(request):
    """
    Returns a list of all locations as dictionaries.
    If a school_id is provided, returns locations linked to that school (school.locations.all()).
    If a lesson_id or pack_id is provided, also returns the selected location (lesson.location or pack.location).
    """
    lesson_id = request.GET.get('lesson_id')
    pack_id = request.GET.get('pack_id')
    school_id = request.GET.get('school_id')

    selected_location = None
    selected_locations = None
    locations_qs = Location.objects.all().values()

    # If school_id provided, get locations from that school; otherwise, get all locations.
    if school_id:
        try:
            school = School.objects.get(pk=school_id)
            selected_locations = [
                {
                    "id": location.id,
                    "name": location.name,
                }
                for location in school.locations.all()
            ]
        except School.DoesNotExist:
            print("Error on school_id")

    # Check for lesson_id or pack_id to set the selected location.
    elif lesson_id:
        try:
            lesson = Lesson.objects.get(pk=lesson_id)
            selected_location_obj = lesson.location  # assuming lesson.location is a Location instance
            selected_location = {"id": selected_location_obj.id, "name": selected_location_obj.name} if selected_location_obj else None
        except Lesson.DoesNotExist:
            selected_location = None
    elif pack_id:
        try:
            pack = Pack.objects.get(pk=pack_id)
            selected_location_obj = pack.location  # assuming pack.location is a Location instance
            selected_location = {"id": selected_location_obj.id, "name": selected_location_obj.name} if selected_location_obj else None
        except Pack.DoesNotExist:
            selected_location = None

    response_data = {
        "locations": list(locations_qs),
    }
    if selected_location:
        response_data["selected_location"] = selected_location
    if selected_locations:
        response_data["selected_locations"] = selected_locations

    return Response(response_data, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_school_subjects(request):
    """
    Updates the subjects associated with a school.
    Expects a JSON payload:
      {
         "school_id": <school id>,
         "subject_ids": [list of subject ids]
      }
    The school's subjects will be replaced with the provided subjects.
    """
    try:
        data = request.data
        school_id = data.get("school_id")
        subject_ids = data.get("subject_ids")
        if not school_id or subject_ids is None:
            return Response({"success": False, "error": "Missing school_id or subject_ids"}, status=400)
        school = School.objects.get(pk=school_id)
        subjects = Sport.objects.filter(id__in=subject_ids)
        school.sports.set(subjects)
        school.save()
        return Response({"success": True, "school_id": school.id}, status=200)
    except School.DoesNotExist:
        return Response({"success": False, "error": "School not found"}, status=404)
    except Exception as e:
        return Response({"success": False, "error": str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_school_locations(request):
    """
    Updates the subjects associated with a school.
    Expects a JSON payload:
      {
         "school_id": <school id>,
         "subject_ids": [list of subject ids]
      }
    The school's subjects will be replaced with the provided subjects.
    """
    try:
        data = request.data
        school_id = data.get("school_id")
        location_ids = data.get("location_ids")
        if not school_id or location_ids is None:
            return Response({"success": False, "error": "Missing school_id or location_ids"}, status=400)
        school = School.objects.get(pk=school_id)
        locations = Location.objects.filter(id__in=location_ids)
        school.locations.set(locations)
        school.save()
        return Response({"success": True, "school_id": school.id}, status=200)
    except School.DoesNotExist:
        return Response({"success": False, "error": "School not found"}, status=404)
    except Exception as e:
        return Response({"success": False, "error": str(e)}, status=500)
    
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_location(request):
    """
    Add the Location for a lesson, pack or school.
    """
    lesson_id = request.data.get("lesson_id")
    pack_id = request.data.get("pack_id")
    school_id = request.data.get("school_id")
    location_name = request.data.get("location_name")
    location_address = request.data.get("location_address")
    
    if not location_name:
        return Response({"error": "É necessário fornecer location_name"}, status=400)
    if not location_address:
        return Response({"error": "É necessário fornecer location_address"}, status=400)
    if not lesson_id and not pack_id and not school_id:
        return Response({"error": "É necessário fornecer lesson_id ou pack_id ou school_id"}, status=400)
        
    lesson = None
    pack = None
    school = None
    
    location = Location.objects.create(name=location_name, address=location_address)
    
    if lesson_id:
        lesson = get_object_or_404(Lesson, id=lesson_id)
        lesson.school.locations.add(location)
        lesson.location = location
        lesson.save()
    elif pack_id:
        pack = get_object_or_404(Pack, id=pack_id)
        pack.school.locations.add(location)
        if pack.type == "private":
            for lesson in pack.lessons.all():
                lesson.location = location
                lesson.save()
        pack.location = location
        pack.save()
    elif school_id:
        school = get_object_or_404(School, id=school_id)
        school.locations.add(location)
        
    status_msg = "location set"
    
    return Response({"status": status_msg}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_subject(request):
    """
    Add the Subject for a lesson, pack or school.
    """
    print(f"Incoming payload:\n{request.data}")
    lesson_id = request.data.get("lesson_id")
    pack_id = request.data.get("pack_id")
    school_id = request.data.get("school_id")
    subject_name = request.data.get("subject_name")
    
    if not subject_name:
        return Response({"error": "É necessário fornecer subject_name"}, status=400)
    if not lesson_id and not pack_id and not school_id:
        return Response({"error": "É necessário fornecer lesson_id ou pack_id ou school_id"}, status=400)
        
    lesson = None
    pack = None
    school = None
    
    subject = Sport.objects.create(name=subject_name)
    if not subject:
        return Response({"error": "Subject was not created"}, status=400)
    else:
        print(f"Subject created: {subject}")
    
    if lesson_id:
        lesson = get_object_or_404(Lesson, id=lesson_id)
        lesson.school.sports.add(subject)
        lesson.sport = subject
        lesson.save()
    elif pack_id:
        pack = get_object_or_404(Pack, id=pack_id)
        pack.school.sports.add(subject)
        if pack.type == "private":
            for lesson in pack.lessons.all():
                lesson.sport = subject
                lesson.save()
        pack.sport = subject
        pack.save()
    elif school_id:
        school = get_object_or_404(School, id=school_id)
        school.sports.add(subject)
        
    status_msg = "Subject set"
    
    return Response({"status": status_msg}, status=status.HTTP_200_OK)


@csrf_exempt  # Consider using proper CSRF handling or a DRF-based view if needed.
def create_review(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse(
                {'error': 'Invalid JSON.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        rating = data.get('rating')
        description = data.get('description', '')
        if rating is None:
            return JsonResponse(
                {'error': 'Rating is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            rating = float(rating)
        except (ValueError, TypeError):
            return JsonResponse(
                {'error': 'Invalid rating provided.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create the review using the authenticated user.
        review = Review.objects.create(
            user=request.user,
            rating=rating,
            description=description
        )
        return JsonResponse(
            {'message': 'Review created successfully', 'review_id': review.id},
            status=status.HTTP_201_CREATED
        )
    return JsonResponse({'error': 'Method not allowed.'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


def parse_date(value):
    # Handle None, NaN, and pandas NaT
    if value is None or pd.isna(value):
        return None
    # pandas Timestamp or datetime
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    try:
        return dt.date.fromisoformat(str(value))
    except Exception:
        return None


def parse_time(value):
    # Handle None, NaN, and pandas NaT
    if value is None or pd.isna(value):
        return None
    if isinstance(value, dt.datetime):
        return value.time()
    if isinstance(value, dt.time):
        return value
    try:
        return dt.time.fromisoformat(str(value))
    except Exception:
        return None

class BulkImportView(views.APIView):
    """
    Endpoint to handle bulk Excel import for Students, Lessons, Packs, and Payments.
    Only Admin users may import. Imported objects are linked to the user's current school.
    """
    serializer_class = CSVUploadSerializer

    def post(self, request, *args, **kwargs):
        if getattr(request.user, 'current_role', None) != 'Admin':
            return Response({'detail': 'Admin privileges required.'},
                            status=status.HTTP_403_FORBIDDEN)

        try:
            school = School.objects.get(pk=request.user.current_school_id)
        except School.DoesNotExist:
            return Response({'detail': 'Invalid school.'},
                            status=status.HTTP_400_BAD_REQUEST)

        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        excel_file = serializer.validated_data['file']

        # load & clean sheets
        xl = pd.ExcelFile(excel_file)
        raw = {}
        for name in xl.sheet_names:
            df = xl.parse(name, header=1).astype(object)
            df = df.where(pd.notna(df), None)
            df.columns = [col.lower() for col in df.columns]
            if 'id' in df.columns and 'old_id_str' not in df.columns:
                df.rename(columns={'id': 'old_id_str'}, inplace=True)
            raw[name.lower()] = df

        results = {}
        for key in ('students','lessons','packs','payments'):
            if key in raw:
                fn = getattr(self, f'_import_{key}')
                results[key] = fn(raw.pop(key), school)

        for name, df in raw.items():
            fn = getattr(self, f'_import_{name.rstrip("s")}', None)
            results[name] = fn(df, school) if fn else {'error': 'No importer for this sheet.'}

        return Response(results, status=status.HTTP_200_OK)

    def _parse_decimal(self, value):
        """Parse a value into a Decimal, defaulting to 0.00 on failure or empty, filtering out NaNs."""
        if value is None:
            return Decimal('0.00')
        # handle float NaN
        if isinstance(value, float) and math.isnan(value):
            return Decimal('0.00')
        # handle string representations
        if isinstance(value, str):
            if not value.strip() or value.strip().lower() == 'nan':
                return Decimal('0.00')
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return Decimal('0.00')

    def _parse_int(self, value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _parse_bool(self, value):
        if isinstance(value, bool): return value
        val = str(value).strip().lower()
        return val in ('true', '1', 'yes', 'y')

    def _import_student(self, df: pd.DataFrame, school: School):
        successes, errors = [], []
        for idx, row in df.iterrows():
            old_id = str(row.get('old_id_str') or '')
            try:
                student, created = Student.objects.update_or_create(
                    old_id_str=old_id,
                    defaults={
                        'first_name': row.get('first_name') or '',
                        'last_name':  row.get('last_name') or '',
                        'birthday':   parse_date(row.get('birthday')),
                        'level':      self._parse_int(row.get('level')),
                    }
                )
                school.students.add(student)
                successes.append({'row': idx+2, 'old_id': old_id, 'new_id': student.id, 'created': created})
            except Exception as e:
                errors.append({'row': idx+2, 'old_id': old_id, 'error': str(e)})
        return {'imported': len(successes), 'errors': errors}

    def _import_instructor(self, df: pd.DataFrame, school: School):
        successes, errors = [], []
        for idx, row in df.iterrows():
            old_id = str(row.get('old_id_str') or '')
            try:
                instructor, created = Instructor.objects.update_or_create(
                    old_id_str=old_id,
                    defaults={
                        'first_name': row.get('first_name') or '',
                        'last_name':  row.get('last_name') or '',

                    }
                )
                school.instructors.add(instructor)
                successes.append({'row': idx+2, 'old_id': old_id, 'new_id': instructor.id, 'created': created})
            except Exception as e:
                errors.append({'row': idx+2, 'old_id': old_id, 'error': str(e)})
        return {'imported': len(successes), 'errors': errors}

    def _import_lesson(self, df: pd.DataFrame, school: School):
        successes, errors = [], []
        for idx, row in df.iterrows():
            old_id = str(row.get('old_id_str') or '')
            try:
                lesson, created = self._create_or_update_lesson(row, old_id)
                school.lessons.add(lesson)
                successes.append({'row': idx+2, 'lesson_id': lesson.id, 'created': created})
            except Exception as e:
                errors.append({'row': idx+2, 'error': str(e)})
        return {'imported': len(successes), 'errors': errors}

    def _create_or_update_lesson(self, row, old_id):
        lesson_date = parse_date(row.get('date'))
        lesson_time = parse_time(row.get('start_time'))
        duration    = self._parse_int(row.get('duration_in_minutes'))
        if duration is None and row.get('type') == 'group':
            duration = 60
        duration = duration or 60
        class_no    = self._parse_int(row.get('class_number'))
        price       = self._parse_decimal(row.get('price'))
        lesson, created = Lesson.objects.update_or_create(
            old_id_str=old_id,
            defaults={
                'date': lesson_date,
                'start_time': lesson_time,
                'duration_in_minutes': duration,
                'class_number': class_no,
                'price': price,
            }
        )
        if row.get('student_ids'):
            ids = [s.strip() for s in str(row['student_ids']).split(',') if s.strip()]
            lesson.students.set(Student.objects.filter(old_id_str__in=ids))
        if row.get('instructor_ids'):
            ids = [s.strip() for s in str(row['instructor_ids']).split(',') if s.strip()]
            lesson.instructors.set(Instructor.objects.filter(old_id_str__in=ids))
        if row.get('pack_ids'):
            ids = [s.strip() for s in str(row['pack_ids']).split(',') if s.strip()]
            lesson.packs.set(Pack.objects.filter(old_id_str__in=ids))
        return lesson, created

    def _import_pack(self, df, school):
        successes, errors = [], []
        for idx, row in df.iterrows():
            old_id = str(row.get('old_id_str') or '')
            try:
                raw_date = row.get('date')
                raw_finished = row.get('finished_date')
                if raw_date and str(raw_date).strip().lower() != 'nan':
                    pack_date = parse_date(raw_date)
                elif raw_finished and str(raw_finished).strip().lower() != 'nan':
                    pack_date = parse_date(raw_finished)
                else:
                    pack_date = dt.date.today()
                defaults = {
                    'date':                pack_date,
                    'type':                row.get('type') or 'private',
                    'number_of_classes':   self._parse_int(row.get('number_of_classes')) or 0,
                    'number_of_classes_left': self._parse_int(row.get('number_of_classes_left'))
                                              or self._parse_int(row.get('number_of_classes')) or 0,
                    'duration_in_minutes': self._parse_int(row.get('duration_in_minutes')) or 0,
                    'price':               self._parse_decimal(row.get('price')),
                    'is_done':             self._parse_bool(row.get('is_done')),
                    'is_paid':             self._parse_bool(row.get('is_paid')),
                    'is_suspended':        self._parse_bool(row.get('is_suspended')),
                    'debt':                self._parse_decimal(row.get('debt')),
                    'finished_date':       parse_date(row.get('finished_date'))
                    if row.get('finished_date') else None,
                    'expiration_date':     parse_date(row.get('expiration_date'))
                    if row.get('expiration_date') else None,
                }
                pack, created = Pack.objects.update_or_create(
                    old_id_str=old_id,
                    defaults=defaults
                )
                school.packs.add(pack)
                # M2M relationships
                if row.get('student_ids'):
                    sids = [s.strip() for s in str(row['student_ids']).split(',') if s.strip()]
                    pack.students.set(Student.objects.filter(old_id_str__in=sids))
                if row.get('lesson_ids'):
                    sids = [s.strip() for s in str(row['lesson_ids']).split(',') if s.strip()]
                    pack.lessons.set(Lesson.objects.filter(old_id_str__in=sids))
                if row.get('instructor_ids'):
                    iids = [s.strip() for s in str(row['instructor_ids']).split(',') if s.strip()]
                    pack.instructors.set(Instructor.objects.filter(old_id_str__in=iids))
                if row.get('parent_ids'):
                    pids = [s.strip() for s in str(row['parent_ids']).split(',') if s.strip()]
                    pack.parents.set(UserAccount.objects.filter(id__in=pids))
                if row.get('sport_id'):
                    try:
                        sport = Sport.objects.get(pk=self._parse_int(row.get('sport_id')))
                        pack.sport = sport
                        pack.save()
                    except Sport.DoesNotExist:
                        pass
                successes.append({'row': idx+2, 'new_id': pack.id, 'created': created})
            except Exception as e:
                errors.append({'row': idx+2, 'old_id': old_id, 'error': str(e)})
        return {'imported': len(successes), 'errors': errors}

    def _import_payment(self, df, school):
        successes, errors = [], []
        for idx, row in df.iterrows():
            old_id = str(row.get('old_id_str') or '')
            try:
                uid = self._parse_int(row.get('user_id'))
                user = UserAccount.objects.filter(pk=uid).first() if uid else None
                desc = row.get('description') if row.get('description') is not None else {}
                payment, created = Payment.objects.update_or_create(
                    old_id_str=old_id,
                    defaults={
                        'value': self._parse_decimal(row.get('value')),
                        'user': user,
                        'description': desc,
                    }
                )
                school.payments.add(payment)
                if row.get('pack_ids'):
                    pids = [s.strip() for s in str(row['pack_ids']).split(',') if s.strip()]
                    payment.packs.set(Pack.objects.filter(old_id_str__in=pids))
                if row.get('lesson_ids'):
                    lids = [s.strip() for s in str(row['lesson_ids']).split(',') if s.strip()]
                    payment.lessons.set(Lesson.objects.filter(old_id_str__in=lids))
                successes.append({'row': idx+2, 'payment_id': payment.id, 'created': created})
            except Exception as e:
                errors.append({'row': idx+2, 'error': str(e)})
        return {'imported': len(successes), 'errors': errors}



class ExcelTemplateView(views.APIView):
    """
    Returns a single .xlsx file with separate sheets for
    Students, Packs, Lessons, and Payments.
    Includes placeholders, data validations, and live conditional formatting
    to highlight invalid cells (red), incomplete rows (orange), and ready rows (green).
    """
    def get(self, request, *args, **kwargs):
        # Column definitions
        templates = {
            'Student': ['first_name', 'last_name', 'birthday', 'level'],
            'Pack': [
                'date', 'type', 'number_of_classes', 'number_of_classes_left',
                'duration_in_minutes', 'price', 'debt', 'is_paid',
                'is_done', 'is_suspended', 'expiration_date',
                'student_ids', 'instructor_ids'
            ],
            'Lesson': ['date', 'start_time', 'duration_in_minutes',
                       'class_number', 'price', 'type', 'student_ids'],
            'Payment': ['value', 'user_id', 'pack_ids', 'lesson_ids'],
        }

        # Fields required for a row to be complete
        required = {
            'Student': ['first_name', 'last_name', 'birthday'],
            'Pack': ['date', 'type', 'number_of_classes', 'duration_in_minutes', 'price'],
            'Lesson': ['date', 'start_time', 'duration_in_minutes'],
            'Payment': ['value', 'user_id'],
        }

        # Placeholder hints
        metadata = {
            'first_name': 'Text (e.g. John)',
            'last_name': 'Text (e.g. Doe)',
            'birthday': 'Date YYYY-MM-DD',
            'level': 'Integer (e.g. 1)',
            'date': 'Date YYYY-MM-DD',
            'type': 'Text (e.g. private)',
            'number_of_classes': 'Integer',
            'number_of_classes_left': 'Integer',
            'duration_in_minutes': 'Integer',
            'price': 'Decimal (e.g. 100.00)',
            'debt': 'Decimal or 0',
            'is_paid': 'TRUE or FALSE',
            'is_done': 'TRUE or FALSE',
            'is_suspended': 'TRUE or FALSE',
            'expiration_date': 'Date YYYY-MM-DD or blank',
            'student_ids': 'CSV of Student IDs',
            'instructor_ids': 'CSV of Instructor IDs',
            'start_time': 'Time HH:MM:SS',
            'class_number': 'Integer or blank',
            'value': 'Decimal (e.g. 50.00)',
            'user_id': 'Integer User PK',
            'pack_ids': 'CSV of Pack IDs',
            'lesson_ids': 'CSV of Lesson IDs',
        }

        # Styles (ARGB)
        placeholder_font = Font(italic=True, color="FF888888")
        placeholder_fill = PatternFill(
            fill_type="solid",
            start_color="FFDDDDDD",
            end_color="FFDDDDDD",
        )
        error_fill = PatternFill(
            fill_type="solid",
            start_color="FFFFC7CE",
            end_color="FFFFC7CE",
        )
        warning_fill = PatternFill(
            fill_type="solid",
            start_color="FFFFEB9C",
            end_color="FFFFEB9C",
        )
        success_fill = PatternFill(
            fill_type="solid",
            start_color="FFC6EFCE",
            end_color="FFC6EFCE",
        )

        wb = Workbook()

        # Help sheet with color key
        help_ws = wb.active
        help_ws.title = 'Help'
        help_ws.append(['Instructions:'])
        help_ws.append(['• Don’t remove or rename sheets.'])
        help_ws.append(['• Row 1 = field name; Row 2 = placeholder.'])
        help_ws.append(['• Enter data from row 3 onward.'])
        help_ws.append(['• Red = invalid cell; Orange = incomplete row; Green = row complete.'])
        help_ws.column_dimensions['A'].width = 25
        help_ws.column_dimensions['B'].width = 50

        # Add a color key for debugging
        # Row 7: red
        cell = help_ws.cell(row=7, column=1, value='')
        cell.fill = error_fill
        help_ws.cell(row=7, column=2, value='Invalid cell (should appear red)')
        # Row 8: orange
        cell = help_ws.cell(row=8, column=1, value='')
        cell.fill = warning_fill
        help_ws.cell(row=8, column=2, value='Incomplete row cell (should appear orange)')
        # Row 9: green
        cell = help_ws.cell(row=9, column=1, value='')
        cell.fill = success_fill
        help_ws.cell(row=9, column=2, value='Complete row cell (should appear green)')

        for sheet_name, cols in templates.items():
            ws = wb.create_sheet(title=sheet_name)
            col_letters = {}

            # Header + placeholder row
            for idx, field in enumerate(cols, start=1):
                hdr = ws.cell(row=1, column=idx, value=field)
                letter = hdr.column_letter
                col_letters[field] = letter

                ph = ws.cell(row=2, column=idx, value=metadata.get(field, ''))
                ph.font = placeholder_font
                ph.fill = placeholder_fill

                ws.column_dimensions[letter].width = 25

            max_row = 5000

            # Cell-level validation + red invalid formatting
            for field, letter in col_letters.items():
                if field.endswith('_date') or field == 'date':
                    test = f"OR(ISBLANK({letter}3), ISNUMBER({letter}3))"
                    dv = DataValidation(
                        type='custom',
                        formula1=test,
                        allow_blank=True
                    )
                    dv.errorTitle = 'Invalid Date'
                    dv.error = 'Must be a valid date'
                    dv.prompt = 'YYYY-MM-DD'

                elif 'time' in field:
                    test = f"OR(ISBLANK({letter}3), ISNUMBER({letter}3))"
                    dv = DataValidation(
                        type='custom',
                        formula1=test,
                        allow_blank=True
                    )
                    dv.errorTitle = 'Invalid Time'
                    dv.error = 'Must be a valid time'
                    dv.prompt = 'HH:MM:SS'

                elif field in ['level', 'number_of_classes', 'number_of_classes_left', 'duration_in_minutes', 'class_number']:
                    test = f"OR(ISBLANK({letter}3), AND(ISNUMBER({letter}3), {letter}3>=0))"
                    dv = DataValidation(
                        type='whole',
                        operator='greaterThanOrEqual',
                        formula1='0',
                        allow_blank=True
                    )
                    dv.errorTitle = 'Invalid Number'
                    dv.error = 'Enter a non-negative integer'

                elif field in ['price', 'debt', 'value']:
                    test = f"OR(ISBLANK({letter}3), AND(ISNUMBER({letter}3), {letter}3>=0))"
                    dv = DataValidation(
                        type='decimal',
                        operator='greaterThanOrEqual',
                        formula1='0',
                        allow_blank=True
                    )
                    dv.errorTitle = 'Invalid Amount'
                    dv.error = 'Enter a non-negative number'

                elif field in ['is_paid', 'is_done', 'is_suspended']:
                    test = f"OR(ISBLANK({letter}3), {letter}3=\"TRUE\", {letter}3=\"FALSE\")"
                    dv = DataValidation(
                        type='list',
                        formula1='"TRUE,FALSE"',
                        allow_blank=True
                    )
                    dv.errorTitle = 'Invalid Choice'
                    dv.error = 'TRUE or FALSE only'

                else:
                    continue

                dv.add(f"{letter}3:{letter}{max_row}")
                ws.add_data_validation(dv)

                ws.conditional_formatting.add(
                    f"{letter}3:{letter}{max_row}",
                    FormulaRule(formula=[f"NOT({test})"], fill=error_fill)
                )

            # Row-level orange (partial) and green (complete)
            first = next(iter(col_letters.values()))
            last  = next(reversed(col_letters.values()))

            # Build required-field tests
            req_tests = []
            for rf in required[sheet_name]:
                lt = col_letters[rf]
                if rf.endswith('_date') or rf == 'date' or 'time' in rf:
                    req_tests.append(f"AND(NOT(ISBLANK({lt}3)), ISNUMBER({lt}3))")
                elif rf in ['level','number_of_classes','duration_in_minutes','class_number']:
                    req_tests.append(f"AND(NOT(ISBLANK({lt}3)), ISNUMBER({lt}3), {lt}3>=0)")
                elif rf in ['price','debt','value']:
                    req_tests.append(f"AND(NOT(ISBLANK({lt}3)), ISNUMBER({lt}3), {lt}3>=0)")
                elif rf in ['is_paid','is_done','is_suspended']:
                    req_tests.append(f"OR({lt}3=\"TRUE\",{lt}3=\"FALSE\")")
                else:
                    req_tests.append(f"NOT(ISBLANK({lt}3))")

            green_formula  = f"AND({','.join(req_tests)})"
            orange_formula = f"AND(COUNTA({first}3:{last}3)>0, NOT({green_formula}))"

            ws.conditional_formatting.add(
                f"{first}3:{last}{max_row}",
                FormulaRule(formula=[orange_formula], fill=warning_fill)
            )
            ws.conditional_formatting.add(
                f"{first}3:{last}{max_row}",
                FormulaRule(formula=[green_formula], fill=success_fill)
            )

        # Write out and return
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="bulk_import_template.xlsx"'}
        )